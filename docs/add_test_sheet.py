from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = load_workbook("docs/CRM_Issue_Tracker.xlsx")

ws = wb.create_sheet("Test Scenarios")

# ── Styles ─────────────────────────────────────────────────────────────────────
thin = Side(style="thin", color="CCCCCC")
med  = Side(style="medium", color="AAAAAA")
border      = Border(left=thin, right=thin, top=thin, bottom=thin)
thick_border = Border(left=med,  right=med,  top=med,  bottom=med)

HEADER_FILL   = PatternFill("solid", fgColor="1E3A5F")
SECTION_FILL  = PatternFill("solid", fgColor="2E6DA4")
PASS_FILL     = PatternFill("solid", fgColor="E8F5E9")
FAIL_FILL     = PatternFill("solid", fgColor="FFEBEE")
SKIP_FILL     = PatternFill("solid", fgColor="FFF8E1")
ALT_FILL      = PatternFill("solid", fgColor="F5F9FF")
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")

PRIO_FILLS = {
    "Critical": PatternFill("solid", fgColor="FF4444"),
    "High":     PatternFill("solid", fgColor="FF8C00"),
    "Medium":   PatternFill("solid", fgColor="FFD700"),
    "Low":      PatternFill("solid", fgColor="90EE90"),
}

def hf(size=10, color="FFFFFF", bold=True):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def cf(bold=False, color="222222", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def ac(h="left", v="top", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ── Column definitions ─────────────────────────────────────────────────────────
headers = [
    "TC-ID", "Module", "Feature / Scenario",
    "Test Steps", "Expected Result",
    "Type", "Priority", "Status",
    "Related Issue", "Notes"
]
widths = [10, 16, 36, 58, 42, 13, 11, 11, 13, 22]

# ── Title ──────────────────────────────────────────────────────────────────────
ws.merge_cells("A1:J1")
t = ws["A1"]
t.value = "Ascendons CRM — Test Scenarios & QA Checklist"
t.font = Font(bold=True, size=16, color="1E3A5F", name="Calibri")
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 38

ws.merge_cells("A2:J2")
s = ws["A2"]
s.value = "Total Scenarios: 120  |  Modules: 18  |  Types: Functional · Integration · Security · Edge Case · Regression"
s.font = Font(italic=True, size=10, color="555555", name="Calibri")
s.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 18
ws.row_dimensions[3].height = 8

# Header row
for ci, (h, w) in enumerate(zip(headers, widths), 1):
    c = ws.cell(row=4, column=ci, value=h)
    c.fill   = HEADER_FILL
    c.font   = hf()
    c.alignment = ac("center", "center")
    c.border = border
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[4].height = 26

ws.freeze_panes = "A5"
ws.auto_filter.ref = "A4:J4"

# ── Data ───────────────────────────────────────────────────────────────────────
# Format: (tc_id, module, feature, steps, expected, type_, priority, status, related, notes)
# Section separator rows: ("SECTION", title)

data = [

# ════════════════════════════════════════════════════════════════
("SECTION", "AUTH & MULTI-TENANCY"),
# ════════════════════════════════════════════════════════════════
("TC-001","Auth","User Login - Valid Credentials",
 "1. POST /api/v1/auth/login with valid email+password\n2. Inspect response",
 "200 OK with JWT token; token contains tenantId, userId, roles",
 "Functional","Critical","Untested","",""),
("TC-002","Auth","User Login - Invalid Password",
 "1. POST /api/v1/auth/login with wrong password",
 "401 Unauthorized; no token returned; error message shown",
 "Functional","High","Untested","",""),
("TC-003","Auth","JWT Expiry Handling",
 "1. Use an expired JWT token in Authorization header\n2. Call any protected endpoint",
 "401 Unauthorized; frontend redirects to /login",
 "Functional","High","Untested","",""),
("TC-004","Auth","Multi-Tenant Data Isolation",
 "1. Log in as Tenant A user\n2. Call GET /leads\n3. Log in as Tenant B user\n4. Call GET /leads",
 "Each tenant sees ONLY their own data; no cross-tenant leakage",
 "Security","Critical","Untested","CRM-048","Most critical test"),
("TC-005","Auth","Role-Based Access - Forbidden Action",
 "1. Log in as Field Engineer role\n2. Attempt DELETE /leads/{id}",
 "403 Forbidden; action blocked by RBAC",
 "Security","Critical","Untested","CRM-002",""),
("TC-006","Auth","Password Reset Flow",
 "1. POST /auth/forgot-password with email\n2. Check email for reset link\n3. POST /auth/reset-password with token",
 "Reset email sent; password updated; old token invalidated",
 "Functional","High","Untested","CRM-006","Email service must be real"),
("TC-007","Auth","Tenant Registration",
 "1. POST /auth/register with org name, email, password\n2. Verify tenant created in DB",
 "201 Created; tenant record; default roles seeded; welcome email sent",
 "Functional","High","Untested","",""),

# ════════════════════════════════════════════════════════════════
("SECTION", "LEAD MANAGEMENT"),
# ════════════════════════════════════════════════════════════════
("TC-008","Leads","Create Lead - All Required Fields",
 "1. Navigate to /leads/new\n2. Fill all required fields\n3. Submit form",
 "Lead created with auto-generated ID (LEAD-YYYY-MM-XXXXX); appears in list",
 "Functional","High","Untested","",""),
("TC-009","Leads","Create Lead - Missing Required Field",
 "1. Submit lead form without firstName",
 "Form shows validation error; no API call made; lead not created",
 "Functional","High","Untested","CRM-003",""),
("TC-010","Leads","Lead BANT Score Calculation",
 "1. Create a lead with Budget=High, Authority=Decision Maker, Need=Strong, Timeline=Immediate\n2. Check score",
 "BANT score ≥ 80; lead classified as Hot",
 "Functional","Medium","Untested","",""),
("TC-011","Leads","Lead Status Transition",
 "1. Open lead in NEW status\n2. Change status to CONTACTED\n3. Change to QUALIFIED",
 "Status updates saved; history log records each transition with timestamp",
 "Functional","High","Untested","",""),
("TC-012","Leads","Lead Assignment to User",
 "1. Open a lead\n2. Assign to a team member\n3. Team member logs in",
 "Assignee sees lead in their queue; notification sent to assignee",
 "Functional","Medium","Untested","",""),
("TC-013","Leads","Lead Conversion to Opportunity",
 "1. Open a QUALIFIED lead\n2. Click Convert\n3. Fill opportunity details",
 "Opportunity created; lead status set to CONVERTED; linked records visible",
 "Functional","High","Untested","",""),
("TC-014","Leads","Bulk Lead Delete",
 "1. Select 3 leads via checkbox\n2. Click Bulk Delete\n3. Confirm dialog",
 "All 3 leads soft-deleted; removed from list; recoverable from DB",
 "Functional","Medium","Untested","",""),
("TC-015","Leads","Lead Search & Filter",
 "1. Search by company name\n2. Filter by status=NEW\n3. Sort by createdAt DESC",
 "Only matching leads shown; pagination works correctly",
 "Functional","Medium","Untested","",""),

# ════════════════════════════════════════════════════════════════
("SECTION", "SALES PIPELINE & OPPORTUNITIES"),
# ════════════════════════════════════════════════════════════════
("TC-016","Opportunities","Create Opportunity",
 "1. Navigate to /opportunities/new\n2. Fill deal name, value, stage, close date\n3. Submit",
 "Opportunity created; visible in pipeline kanban and list view",
 "Functional","High","Untested","",""),
("TC-017","Opportunities","Move Opportunity Across Pipeline Stages",
 "1. Open opportunity\n2. Change stage from Prospecting → Qualification → Proposal",
 "Stage transitions saved; probability auto-updates per stage; history logged",
 "Functional","High","Untested","",""),
("TC-018","Opportunities","Opportunity Win/Loss",
 "1. Set opportunity to Closed Won\n2. Set another to Closed Lost with reason",
 "Won: revenue recorded. Lost: reason saved. Both removed from active pipeline.",
 "Functional","Medium","Untested","",""),

# ════════════════════════════════════════════════════════════════
("SECTION", "ASSETS & EQUIPMENT"),
# ════════════════════════════════════════════════════════════════
("TC-019","Assets","Create Asset",
 "1. Navigate to /assets/new\n2. Fill serial no, model, brand, category, account\n3. Submit",
 "Asset created with auto code (ASSET-YYYY-MM-XXXXX); visible in /assets",
 "Functional","High","Untested","",""),
("TC-020","Assets","Duplicate Serial Number Rejection",
 "1. Create asset with serialNo=SN001\n2. Try creating another with same serialNo",
 "409 Conflict returned; second asset not created; error shown in UI",
 "Functional","High","Untested","",""),
("TC-021","Assets","Asset Warranty Expiry Alert",
 "1. Create asset with warrantyExpiry = today + 25 days\n2. Run AssetWarrantyExpiryScheduler manually",
 "Log entry shows 30-day warning for the asset; notification sent (once wired)",
 "Functional","Medium","Untested","CRM-012",""),
("TC-022","Assets","Filter Assets by Account",
 "1. Navigate to /assets?accountId=ACC-001",
 "Only assets linked to that account returned; others excluded",
 "Functional","Medium","Untested","",""),
("TC-023","Assets","Asset Status Update",
 "1. Open asset detail\n2. Change status from ACTIVE to UNDER_REPAIR",
 "Status updated; reflected in list view with correct badge color",
 "Functional","Low","Untested","",""),

# ════════════════════════════════════════════════════════════════
("SECTION", "AMC / CONTRACT MANAGEMENT"),
# ════════════════════════════════════════════════════════════════
("TC-024","Contracts","Create AMC Contract",
 "1. Navigate to /contracts/new\n2. Select type=AMC, account, assets, SLA config\n3. Submit",
 "Contract created with number CON-XXXXX; visits auto-generated per visitFrequency",
 "Functional","High","Untested","",""),
("TC-025","Contracts","Activate Contract",
 "1. Open DRAFT contract\n2. Click Activate",
 "Status changes to ACTIVE; start date recorded; SLA clock begins",
 "Functional","High","Untested","",""),
("TC-026","Contracts","Contract Visit Auto-WO Creation",
 "1. Create contract with visit scheduled for today\n2. Trigger ContractVisitScheduler",
 "Work order of type AMC_VISIT auto-created; linked to visit record",
 "Integration","High","Untested","CRM-048","Check tenant isolation"),
("TC-027","Contracts","Contract Renewal",
 "1. Open EXPIRED contract\n2. Click Renew with new dates",
 "New contract created; old one marked RENEWED; visits regenerated",
 "Functional","Medium","Untested","",""),
("TC-028","Contracts","Contract SLA Config Feed to Work Order",
 "1. Contract has SLA responseHrs=4, resolutionHrs=8\n2. Create WO from this contract",
 "WO.slaDeadline = createdAt + 8h; SLA breach triggers if exceeded",
 "Integration","High","Untested","",""),

# ════════════════════════════════════════════════════════════════
("SECTION", "WORK ORDER MANAGEMENT"),
# ════════════════════════════════════════════════════════════════
("TC-029","Work Orders","Create Work Order",
 "1. Navigate to /work-orders/new\n2. Fill type, priority, asset, account\n3. Submit",
 "WO created with number WO-XXXXX; status=OPEN; SLA deadline calculated",
 "Functional","Critical","Untested","",""),
("TC-030","Work Orders","Work Order State Machine - Full Flow",
 "1. Create WO (OPEN)\n2. Assign engineer (ASSIGNED)\n3. Engineer marks En Route\n4. Arrive on site (ON_SITE)\n5. Start checklist (IN_PROGRESS)\n6. Complete checklist (COMPLETED)",
 "Each status transition succeeds; invalid transitions (e.g. OPEN→COMPLETED) rejected with 400",
 "Functional","Critical","Untested","",""),
("TC-031","Work Orders","SLA Breach Detection",
 "1. Create WO with slaDeadline = now - 1 hour\n2. Run SlaBreachScheduler",
 "WO.slaBreached=true; escalation rule fires if configured",
 "Functional","High","Untested","",""),
("TC-032","Work Orders","Assign Engineers to WO",
 "1. Open WO in OPEN status\n2. Assign 2 engineers",
 "Status changes to ASSIGNED; both engineers see WO in their queue",
 "Functional","High","Untested","",""),
("TC-033","Work Orders","Reopen Completed Work Order",
 "1. WO in COMPLETED status\n2. Click Reopen",
 "Status = REOPENED; reopenCount incremented; new SLA deadline set",
 "Functional","Medium","Untested","",""),
("TC-034","Work Orders","Parts Deduction on WO Closure",
 "1. Add part (qty=2) to WO.partsUsed\n2. Close WO",
 "Product.stockQuantity reduced by 2; StockTransaction created with type=PRODUCTION_OUT",
 "Integration","High","Untested","",""),
("TC-035","Work Orders","Insufficient Stock on WO Closure",
 "1. Add part with qty > available stock to WO\n2. Attempt to close WO",
 "400 Bad Request: 'Insufficient stock'; WO stays in IN_PROGRESS; stock unchanged",
 "Edge Case","High","Untested","",""),
("TC-036","Work Orders","Checklist Mandatory Item Fail Blocks Completion",
 "1. Start checklist with mandatory item\n2. Mark item as FAIL with failureAction=Block\n3. Try to complete checklist",
 "Completion blocked; error message shown; WO cannot advance to COMPLETED",
 "Functional","High","Untested","",""),
("TC-037","Work Orders","Kanban View Displays Correct Columns",
 "1. Navigate to /work-orders\n2. Switch to Kanban view",
 "WOs grouped by status; drag is not required but cards show correct status column",
 "Functional","Medium","Untested","",""),

# ════════════════════════════════════════════════════════════════
("SECTION", "SERVICE REQUESTS"),
# ════════════════════════════════════════════════════════════════
("TC-038","Service Requests","Create Service Request",
 "1. Navigate to /service-requests/new\n2. Fill description, source, priority, account\n3. Submit",
 "SR created with number SR-XXXXX; status=OPEN; SLA clock starts",
 "Functional","High","Untested","",""),
("TC-039","Service Requests","Acknowledge Service Request",
 "1. Open SR in OPEN status\n2. Click Acknowledge",
 "Status = ACKNOWLEDGED; acknowledgedAt timestamp saved",
 "Functional","High","Untested","",""),
("TC-040","Service Requests","Convert SR to Work Order",
 "1. Open ACKNOWLEDGED SR\n2. Click Convert to WO\n3. Fill WO details",
 "WO created and linked to SR; SR status = WO_CREATED; workOrderId populated",
 "Integration","High","Untested","",""),
("TC-041","Service Requests","SR Unacknowledged Escalation",
 "1. Create SR with conditionMinutes=30 escalation rule\n2. Wait / manually trigger evaluator after 30 min",
 "EscalationLog created for SR; L1 users notified",
 "Integration","Medium","Untested","CRM-046",""),

# ════════════════════════════════════════════════════════════════
("SECTION", "DISPATCH & SCHEDULING"),
# ════════════════════════════════════════════════════════════════
("TC-042","Dispatch","View Engineer Schedule for Date",
 "1. Navigate to /dispatch\n2. Select today's date",
 "Engineer schedule tiles appear; availability badges shown (AVAILABLE/ON_JOB/LEAVE)",
 "Functional","High","Untested","",""),
("TC-043","Dispatch","Dispatch Work Order to Engineer",
 "1. Open dispatch board\n2. Click Dispatch on open WO\n3. Select AVAILABLE engineer\n4. Confirm",
 "WO status → ASSIGNED; DispatchAssignment created; EngineerSchedule updated to ON_JOB",
 "Functional","High","Untested","",""),
("TC-044","Dispatch","Cannot Dispatch to ON_JOB Engineer",
 "1. Try to dispatch WO to engineer with availability=ON_JOB",
 "Warning shown; dispatch either blocked or flagged as overload",
 "Edge Case","Medium","Untested","",""),
("TC-045","Dispatch","Reassign Work Order",
 "1. WO assigned to Engineer A\n2. Reassign to Engineer B with reason",
 "Old engineer's schedule slot removed; new engineer assigned; audit log entry created",
 "Functional","Medium","Untested","",""),

# ════════════════════════════════════════════════════════════════
("SECTION", "GEO-TRACKING"),
# ════════════════════════════════════════════════════════════════
("TC-046","Geo-Tracking","Engineer Location Update",
 "1. POST /geo/location with lat, lng, workOrderId\n2. GET /geo/locations",
 "Location saved; TTL set for 24h; latest location returned in GET",
 "Functional","High","Untested","",""),
("TC-047","Geo-Tracking","Geo-Event: Arrived Triggers WO ON_SITE",
 "1. POST /geo/events with eventType=Arrived, workOrderId of EN_ROUTE WO",
 "WO status auto-transitions to ON_SITE; WorkOrderGeoEvent persisted",
 "Integration","High","Untested","",""),
("TC-048","Geo-Tracking","Location TTL Expiry",
 "1. Insert engineer location\n2. Wait 24h (or set TTL to 1s in test)\n3. Query location",
 "Record automatically deleted by MongoDB TTL index",
 "Functional","Low","Untested","",""),

# ════════════════════════════════════════════════════════════════
("SECTION", "SKILL MATRIX"),
# ════════════════════════════════════════════════════════════════
("TC-049","Skill Matrix","Add Skill to Technician",
 "1. Navigate to /skill-matrix\n2. Enter userId\n3. Add Skill: VRF, EXPERT, with cert details",
 "Skill record created; appears in skills table; proficiency badge shows EXPERT",
 "Functional","Medium","Untested","",""),
("TC-050","Skill Matrix","Add Training Record",
 "1. Add training: OEM training, score=90, passed=true",
 "Training record saved; pass badge (green tick) shown in table",
 "Functional","Medium","Untested","",""),
("TC-051","Skill Matrix","Expiring Certification Alert",
 "1. Add skill with expiryDate = today + 25 days\n2. Call GET /skill-matrix/expiring?days=30",
 "Skill returned in expiring list; alert should be triggered",
 "Functional","Medium","Untested","",""),

# ════════════════════════════════════════════════════════════════
("SECTION", "SPARE PARTS & PARTS REQUESTS"),
# ════════════════════════════════════════════════════════════════
("TC-052","Parts Requests","Create Parts Request from WO",
 "1. Open WO in IN_PROGRESS status\n2. Navigate to /work-orders/{id}/parts\n3. Add parts request",
 "PartsRequest created; status=PENDING; WO status may change to PENDING_SPARES",
 "Functional","High","Untested","CRM-045",""),
("TC-053","Parts Requests","Approve Parts Request",
 "1. Open PENDING parts request in /parts-requests\n2. Click Approve",
 "Status = APPROVED; stock reserved; WebSocket push to engineer",
 "Functional","High","Untested","",""),
("TC-054","Parts Requests","Reject Parts Request with Reason",
 "1. Open PENDING parts request\n2. Click Reject\n3. Enter reason",
 "Status = REJECTED; reason saved; engineer notified",
 "Functional","Medium","Untested","",""),
("TC-055","Parts Requests","Dispatch Parts",
 "1. APPROVED parts request\n2. Click Dispatch",
 "Status = DISPATCHED; dispatchedAt timestamp saved",
 "Functional","Medium","Untested","",""),
("TC-056","Parts Requests","Receive Parts",
 "1. DISPATCHED parts request\n2. Click Mark Received",
 "Status = RECEIVED; receivedAt saved; WO can resume IN_PROGRESS",
 "Functional","Medium","Untested","",""),
("TC-057","Parts","Auto-Reorder PO Created on Low Stock",
 "1. Set product reorderPoint=10, stockQty=12\n2. Close WO consuming 5 units (stockQty → 7)\n3. Check purchase_orders collection",
 "Draft PO auto-created for the product; poNumber starts with PO-AUTO-",
 "Integration","High","Untested","CRM-039","Verify no duplicates"),

# ════════════════════════════════════════════════════════════════
("SECTION", "VENDOR MANAGEMENT"),
# ════════════════════════════════════════════════════════════════
("TC-058","Vendors","Create Vendor",
 "1. Navigate to /vendors/new\n2. Fill companyName, contactPerson, email, GSTIN\n3. Submit",
 "Vendor created; vendorCode auto-generated (VEN-XXXXX); default rating=3",
 "Functional","High","Untested","",""),
("TC-059","Vendors","Update Vendor Rating",
 "1. Open vendor detail\n2. Click 4-star rating\n3. Confirm update",
 "POST /vendors/{id}/rating with rating=4; vendor.rating updated; badge reflects change",
 "Functional","Medium","Untested","",""),
("TC-060","Vendors","Blacklist Vendor",
 "1. Change vendor status to BLACKLISTED",
 "Vendor marked BLACKLISTED; badge shows red; should be excluded from RFQ auto-suggestions",
 "Functional","Medium","Untested","",""),
("TC-061","Vendors","Vendor Detail Shows Rate Contracts",
 "1. Create rate contract for vendor\n2. Open vendor detail page",
 "Rate contracts section on vendor detail page shows associated contracts",
 "Functional","Low","Untested","CRM-050",""),

# ════════════════════════════════════════════════════════════════
("SECTION", "PROCUREMENT (RFQ / GRN / RATE CONTRACTS / PO)"),
# ════════════════════════════════════════════════════════════════
("TC-062","RFQ","Create RFQ",
 "1. Navigate to /procurement/rfq\n2. Click Create\n3. Add items, select vendors, set deadline",
 "RFQ created with status=OPEN; RFQ number auto-generated",
 "Functional","High","Untested","",""),
("TC-063","RFQ","Submit Vendor Response",
 "1. Open RFQ\n2. POST /procurement/rfq/{id}/vendor-response with vendorId, unitPrice, deliveryDays",
 "Vendor response saved under RFQ; visible in comparison table on RFQ detail page",
 "Functional","High","Untested","",""),
("TC-064","RFQ","Select Vendor Closes RFQ",
 "1. Open RFQ with at least one vendor response\n2. Click Select Vendor",
 "RFQ status = CLOSED; selectedVendorId populated; other vendors notified (future)",
 "Functional","High","Untested","",""),
("TC-065","GRN","Create GRN Against PO",
 "1. Navigate to /procurement/grn\n2. Create GRN: poId, receivedDate, lineItems, qualityStatus",
 "GRN created with number GRN-XXXXX; linked to PO; PO status may update to RECEIVING",
 "Functional","High","Untested","",""),
("TC-066","GRN","GRN Quality Status: Partially Accepted",
 "1. Create GRN with receivedQty < orderedQty for one item\n2. Set qualityStatus=PARTIALLY_ACCEPTED",
 "GRN saved; PO remains in RECEIVING state; remaining qty tracked",
 "Edge Case","Medium","Untested","",""),
("TC-067","Rate Contracts","Create Rate Contract",
 "1. Navigate to /procurement/rate-contracts\n2. Create with vendorId, items, validFrom, validTo",
 "Rate contract created with RC number; status=ACTIVE; appears in active contracts list",
 "Functional","High","Untested","",""),
("TC-068","Rate Contracts","Terminate Rate Contract",
 "1. Open active rate contract\n2. Click Terminate",
 "Status = TERMINATED; no longer appears in active list",
 "Functional","Medium","Untested","",""),
("TC-069","PO Approval","Submit PO for Approval - Amount < 50k",
 "1. Create PO with totalAmount=30000\n2. Click Submit for Approval",
 "approvalWorkflow has 1 step (L1 only); status = SUBMITTED",
 "Functional","High","Untested","CRM-024",""),
("TC-070","PO Approval","Submit PO for Approval - Amount 50k-5L",
 "1. Create PO with totalAmount=200000\n2. Submit for Approval",
 "approvalWorkflow has 2 steps (L1 + L2); status = SUBMITTED",
 "Functional","High","Untested","CRM-024",""),
("TC-071","PO Approval","Submit PO for Approval - Amount > 5L",
 "1. Create PO with totalAmount=600000\n2. Submit for Approval",
 "approvalWorkflow has 3 steps (L1 + L2 + L3); status = SUBMITTED",
 "Functional","High","Untested","CRM-024",""),
("TC-072","PO Approval","L1 Approve → L2 Required Next",
 "1. SUBMITTED PO (L1+L2 workflow)\n2. L1 approver calls POST /approve?level=L1",
 "L1 step status = Approved; PO still SUBMITTED; L2 pending",
 "Functional","High","Untested","CRM-024",""),
("TC-073","PO Approval","Skip L1 and Approve at L2 - Should Fail",
 "1. SUBMITTED PO (L1+L2 workflow)\n2. Try to approve at L2 without L1 approval",
 "400 Bad Request: Level L1 approval required first",
 "Edge Case","High","Untested","CRM-024",""),
("TC-074","PO Approval","Reject PO at Any Level",
 "1. SUBMITTED PO\n2. POST /reject?level=L1 with reason",
 "PO status = CANCELLED; rejectionReason saved; step status = Rejected",
 "Functional","Medium","Untested","CRM-024",""),

# ════════════════════════════════════════════════════════════════
("SECTION", "DEALER MANAGEMENT"),
# ════════════════════════════════════════════════════════════════
("TC-075","Dealers","Create Dealer",
 "1. Navigate to /dealers/new\n2. Fill companyName, tier=GOLD, creditLimit=500000\n3. Submit",
 "Dealer created with dealerCode DLR-XXXXX; status=ACTIVE; credit utilization=0%",
 "Functional","High","Untested","",""),
("TC-076","Dealers","Place Dealer Order Within Credit Limit",
 "1. Dealer has creditLimit=100000, currentCreditUsed=0\n2. Place order totalValue=50000",
 "Order created; currentCreditUsed → 50000; credit utilization shows 50%",
 "Functional","High","Untested","",""),
("TC-077","Dealers","Place Order Exceeding Credit Limit",
 "1. Dealer creditLimit=100000, currentCreditUsed=80000\n2. Try placing order totalValue=30000",
 "400 Bad Request: credit limit exceeded; order not created; UI shows error",
 "Edge Case","Critical","Untested","",""),
("TC-078","Dealers","Suspend Dealer",
 "1. Open active dealer\n2. Click Suspend",
 "Status = SUSPENDED; badge turns red; suspended dealers cannot place orders",
 "Functional","Medium","Untested","",""),
("TC-079","Dealers","Dealer Monthly Performance Aggregation",
 "1. Place 3 orders in current month (2 Delivered, 1 Pending)\n2. Trigger performance aggregation or GET /dealers/{id}/performance",
 "actualSales = sum of Delivered order values; openOrders = 1",
 "Integration","Medium","Untested","CRM-049","Verify tenant isolation"),
("TC-080","Dealers","Dealer Tier Badge Display",
 "1. Create dealers with each tier: PLATINUM, GOLD, SILVER, BRONZE\n2. Open /dealers list",
 "Each dealer shows correct colored badge: purple/amber/gray/orange",
 "Functional","Low","Untested","",""),

# ════════════════════════════════════════════════════════════════
("SECTION", "SERVICE ANALYTICS"),
# ════════════════════════════════════════════════════════════════
("TC-081","Service Analytics","KPI Dashboard Loads",
 "1. Navigate to /analytics/service\n2. Wait for all 3 API calls to complete",
 "All 6 KPI cards populated; WO aging buckets show bar chart; no errors",
 "Functional","High","Untested","",""),
("TC-082","Service Analytics","MTTR Calculation",
 "1. Complete 2 WOs: one took 2h, one took 4h\n2. Check /analytics/service/kpis",
 "mttrHours = 3.0 (average of 2h and 4h)",
 "Functional","High","Untested","",""),
("TC-083","Service Analytics","SLA Compliance Rate",
 "1. Complete 4 WOs: 3 within SLA, 1 breached\n2. Check kpis.slaComplianceRatePct",
 "slaComplianceRatePct = 75.0",
 "Functional","High","Untested","",""),
("TC-084","Service Analytics","First Time Fix Rate",
 "1. Complete 5 WOs: 4 with reopenCount=0, 1 with reopenCount=1\n2. Check kpis.firstTimeFixRatePct",
 "firstTimeFixRatePct = 80.0",
 "Functional","Medium","Untested","",""),
("TC-085","Service Analytics","Volume by Type/Priority/Status",
 "1. Create WOs with different types and priorities\n2. Call /analytics/service/volume",
 "byType, byPriority, byStatus maps contain correct counts",
 "Functional","Medium","Untested","",""),
("TC-086","Service Analytics","Analytics Cache Invalidation",
 "1. Load /analytics/service/kpis (cached)\n2. Complete a new WO\n3. Reload analytics",
 "Cache TTL of 15 min means stale data is acceptable; document expected behaviour",
 "Functional","Low","Untested","",""),

# ════════════════════════════════════════════════════════════════
("SECTION", "INVENTORY ANALYTICS"),
# ════════════════════════════════════════════════════════════════
("TC-087","Inventory Analytics","Dead Stock Report - 90 Day Bucket",
 "1. Product with last StockTransaction > 90 days ago\n2. GET /analytics/service/inventory/dead-stock",
 "Product appears in deadStock90Days list with correct daysSinceLastMovement",
 "Functional","High","Untested","",""),
("TC-088","Inventory Analytics","Reorder Recommendations",
 "1. Product with reorderPoint=20, stockQty=15\n2. GET /analytics/service/inventory/reorder",
 "Product appears in reorder list; suggestedReorderQty = reorderQty or reorderPoint*2",
 "Functional","High","Untested","",""),
("TC-089","Inventory Analytics","Top Consumed Parts",
 "1. Multiple PRODUCTION_OUT transactions for different products\n2. GET /analytics/service/inventory/top-consumed?limit=5",
 "Top 5 products by total consumption returned in descending order",
 "Functional","Medium","Untested","",""),

# ════════════════════════════════════════════════════════════════
("SECTION", "ESCALATION ENGINE"),
# ════════════════════════════════════════════════════════════════
("TC-090","Escalation","Create Escalation Rule",
 "1. Navigate to /admin/settings/escalation\n2. Create rule: SLA_BREACH, conditionMinutes=0, level=L1\n3. Save",
 "Rule created; active=true; appears in rules table",
 "Functional","High","Untested","",""),
("TC-091","Escalation","SLA Breach Triggers Escalation",
 "1. Create WO with slaBreached=true\n2. Run escalation evaluator\n3. Check /escalation/logs",
 "EscalationLog created for the WO; no duplicate log if re-run",
 "Integration","Critical","Untested","",""),
("TC-092","Escalation","No Duplicate Escalation Log",
 "1. WO with slaBreached=true; escalation already fired and unresolved\n2. Run evaluator again",
 "No new EscalationLog created; dedup logic prevents duplicate",
 "Edge Case","High","Untested","",""),
("TC-093","Escalation","Acknowledge Escalation",
 "1. Open open escalation in /admin/settings/escalation\n2. Click Acknowledge",
 "acknowledgedAt and acknowledgedBy populated; button disabled after",
 "Functional","Medium","Untested","",""),
("TC-094","Escalation","Resolve Escalation",
 "1. Open acknowledged escalation\n2. Click Resolve",
 "resolvedAt populated; escalation no longer appears in Open Escalations tab",
 "Functional","Medium","Untested","",""),
("TC-095","Escalation","Inactive Rule Not Evaluated",
 "1. Set escalation rule active=false\n2. Create a WO that would trigger the rule\n3. Run evaluator",
 "No EscalationLog created for inactive rule",
 "Edge Case","Medium","Untested","",""),

# ════════════════════════════════════════════════════════════════
("SECTION", "HR — ATTENDANCE & LEAVE"),
# ════════════════════════════════════════════════════════════════
("TC-096","Attendance","Check In with Valid GPS",
 "1. POST /attendance/check-in with valid lat/lng within office geofence",
 "Attendance record created; checkInTime saved; status=PRESENT",
 "Functional","High","Untested","",""),
("TC-097","Attendance","Check In Outside Geofence",
 "1. POST /attendance/check-in with lat/lng far from any office location",
 "Check-in flagged as IRREGULAR or rejected; spoofDetected handled",
 "Security","High","Untested","CRM-017",""),
("TC-098","Attendance","Check Out",
 "1. Check in at 9:00 AM\n2. Check out at 6:00 PM",
 "totalWorkHours = 9.0; status=PRESENT; OT calculated if > shift hours",
 "Functional","High","Untested","",""),
("TC-099","Leave","Apply for Leave",
 "1. Navigate to /leaves\n2. Apply for 2-day leave with type=CASUAL\n3. Submit",
 "Leave request created with status=PENDING; manager notified",
 "Functional","High","Untested","",""),
("TC-100","Leave","Approve Leave",
 "1. Manager opens leave request\n2. Approves",
 "Status = APPROVED; leave balance deducted; employee notified",
 "Functional","High","Untested","",""),

# ════════════════════════════════════════════════════════════════
("SECTION", "RBAC & PERMISSIONS"),
# ════════════════════════════════════════════════════════════════
("TC-101","RBAC","Field Engineer Profile Permissions",
 "1. Create user with Field Engineer profile\n2. Attempt to access /admin\n3. Attempt to access /work-orders",
 "/admin returns 403; /work-orders returns 200 with own WOs only",
 "Security","Critical","Untested","CRM-005",""),
("TC-102","RBAC","Service Manager Full Field Service Access",
 "1. Create user with Service Manager profile\n2. Access all field service routes",
 "Assets, Contracts, WOs, Dispatch, Skill Matrix, Analytics all accessible",
 "Security","High","Untested","",""),
("TC-103","RBAC","New Tenant Gets Default Profiles",
 "1. Register new tenant\n2. Check profiles collection",
 "8 system profiles cloned for new tenant; all permission objects present including field service modules",
 "Integration","High","Untested","",""),
("TC-104","RBAC","patchMissingPermissions Adds New Modules",
 "1. Existing tenant with old profiles (no DISPATCH permission)\n2. Restart app (runs patchMissingPermissions)\n3. Check profile",
 "DISPATCH, SKILL_MATRIX, VENDORS etc. added to all existing profiles",
 "Regression","High","Untested","",""),

# ════════════════════════════════════════════════════════════════
("SECTION", "PROPOSALS & INVOICES"),
# ════════════════════════════════════════════════════════════════
("TC-105","Proposals","Create Proposal",
 "1. Navigate to /proposals/new\n2. Fill line items, GST, discount\n3. Submit",
 "Proposal created with version 1; total calculated correctly with GST",
 "Functional","High","Untested","",""),
("TC-106","Proposals","Send Proposal to Customer",
 "1. Open DRAFT proposal\n2. Click Send",
 "Status = SENT; sentAt recorded; email notification sent to customer (CRM-007)",
 "Functional","High","Untested","CRM-007","Email must be wired"),
("TC-107","Proposals","Accept Proposal",
 "1. Open SENT proposal\n2. Click Accept",
 "Status = ACCEPTED; opportunity stage updated; invoice can be raised",
 "Functional","High","Untested","CRM-007",""),
("TC-108","Proposals","Proposal Versioning",
 "1. Revise an existing proposal",
 "New version created (v2); old version archived; only latest shown by default",
 "Functional","Medium","Untested","",""),

# ════════════════════════════════════════════════════════════════
("SECTION", "SECURITY & INFRASTRUCTURE"),
# ════════════════════════════════════════════════════════════════
("TC-109","Security","Credentials Not in Source Code",
 "1. Review application.properties in repo\n2. Verify no plaintext credentials",
 "All secrets reference environment variables; no hardcoded passwords or URIs",
 "Security","Critical","Untested","CRM-001","P0 - fix immediately"),
("TC-110","Security","SQL/NoSQL Injection Prevention",
 "1. POST /leads with firstName = '{$where: function(){return true}}'\n2. Inspect DB query",
 "Spring Data MongoDB parameterises queries; injection attempt harmless",
 "Security","High","Untested","",""),
("TC-111","Security","XSS Prevention in Frontend",
 "1. Create lead with name = '<script>alert(1)</script>'\n2. View lead detail page",
 "Script not executed; React escapes output; name shown as literal text",
 "Security","High","Untested","",""),
("TC-112","Security","CORS Policy Enforcement",
 "1. Send request from disallowed origin (http://evil.com)\n2. Check response headers",
 "Request blocked or Access-Control-Allow-Origin header absent/mismatched",
 "Security","High","Untested","CRM-037",""),
("TC-113","Security","Rate Limiting / Brute Force Protection",
 "1. Send 20 failed login attempts from same IP within 1 minute",
 "After threshold, subsequent attempts return 429 Too Many Requests (if implemented)",
 "Security","Medium","Untested","","Not yet implemented - log as feature"),

# ════════════════════════════════════════════════════════════════
("SECTION", "PERFORMANCE & EDGE CASES"),
# ════════════════════════════════════════════════════════════════
("TC-114","Performance","Large Dataset Pagination",
 "1. Seed 1000 leads\n2. GET /leads?page=0&size=20",
 "Response in < 500ms; exactly 20 records; totalElements=1000",
 "Performance","Medium","Untested","",""),
("TC-115","Performance","EscalationService findAll() Under Load",
 "1. Seed 10 tenants × 50 escalation rules each\n2. Trigger evaluateAllRules()",
 "All 500 rules evaluated without memory spike; completes in < 30s",
 "Performance","Medium","Untested","CRM-046",""),
("TC-116","Performance","DealerPerformanceService Aggregation at Scale",
 "1. Seed 5 tenants × 500 orders each\n2. Trigger monthly aggregation",
 "Aggregation completes without OOM; per-tenant data correct",
 "Performance","High","Untested","CRM-049",""),
("TC-117","Edge Case","Soft Delete Excludes Records from All Queries",
 "1. Create and then delete a lead\n2. GET /leads; GET /leads/{id}",
 "Deleted lead absent from list; GET by ID returns 404",
 "Regression","High","Untested","",""),
("TC-118","Edge Case","Concurrent WO Closure - Stock Race Condition",
 "1. Set stock qty=5 for a part\n2. Simultaneously close 2 WOs each consuming 5 units",
 "Only one closure succeeds; the other gets 400 Insufficient Stock; final stock ≥ 0",
 "Edge Case","High","Untested","","Verify with load test"),
("TC-119","Edge Case","PO Auto-Reorder Deduplication",
 "1. Close WO that triggers reorder for Product X\n2. Immediately close another WO using Product X",
 "Only one draft reorder PO created for Product X; second run detects existing draft and skips",
 "Edge Case","Medium","Untested","CRM-039",""),
("TC-120","Regression","Sidebar Navigation Permissions",
 "1. Log in as Field Engineer\n2. Inspect sidebar",
 "Only FIELD_SERVICE and HR sections visible; CRM/Admin/Procurement hidden",
 "Regression","High","Untested","CRM-043",""),
]

# ── Write rows ─────────────────────────────────────────────────────────────────
current_row = 5
tc_count = 0

for item in data:
    if item[0] == "SECTION":
        # Section header row
        ws.merge_cells(f"A{current_row}:J{current_row}")
        c = ws.cell(row=current_row, column=1, value=f"  {item[1]}")
        c.fill   = SECTION_FILL
        c.font   = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = thick_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1
        continue

    tc_id, module, feature, steps, expected, type_, priority, status, related, notes = item
    tc_count += 1

    row_fill = ALT_FILL if tc_count % 2 == 0 else WHITE_FILL

    values = [tc_id, module, feature, steps, expected, type_, priority, status, related, notes]
    for ci, val in enumerate(values, 1):
        c = ws.cell(row=current_row, column=ci, value=val)
        c.fill      = row_fill
        c.font      = cf()
        c.border    = border
        c.alignment = ac()

    # Priority badge
    p_cell = ws.cell(row=current_row, column=7)
    p_key = priority.split()[0] if priority else ""
    if p_key in PRIO_FILLS:
        p_cell.fill = PRIO_FILLS[p_key]
        p_cell.font = Font(bold=True, color="FFFFFF" if p_key in ("Critical","High") else "333333",
                           size=10, name="Calibri")
    p_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Status badge
    s_cell = ws.cell(row=current_row, column=8)
    s_cell.fill = SKIP_FILL
    s_cell.font = cf(bold=True, color="7A6000")
    s_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Type badge
    t_cell = ws.cell(row=current_row, column=6)
    t_cell.alignment = Alignment(horizontal="center", vertical="center")
    if type_ == "Security":
        t_cell.fill = PatternFill("solid", fgColor="FFD0D0")
        t_cell.font = cf(bold=True, color="8B0000")
    elif type_ == "Edge Case":
        t_cell.fill = PatternFill("solid", fgColor="E8D5FF")
        t_cell.font = cf(bold=True, color="4B0082")
    elif type_ == "Integration":
        t_cell.fill = PatternFill("solid", fgColor="D0F0FF")
        t_cell.font = cf(bold=True, color="003366")
    elif type_ == "Performance":
        t_cell.fill = PatternFill("solid", fgColor="FFE8C0")
        t_cell.font = cf(bold=True, color="7A3B00")
    elif type_ == "Regression":
        t_cell.fill = PatternFill("solid", fgColor="C8FFD4")
        t_cell.font = cf(bold=True, color="005A1F")

    ws.row_dimensions[current_row].height = 70
    current_row += 1

# ── Update subtitle with actual count ─────────────────────────────────────────
ws["A2"].value = (
    f"Total Scenarios: {tc_count}  |  Modules: 18  |  "
    "Types: Functional · Integration · Security · Edge Case · Regression · Performance"
)

wb.save("docs/CRM_Issue_Tracker.xlsx")
print(f"Done. Added {tc_count} test scenarios across 18 modules.")
